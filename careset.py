import pandas as pd
import os
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import glob
import warnings
import time

start_time =time.time()

warnings.simplefilter("ignore")


print("Chargement des fichiers tango")
#definition des inputs

chemin=r"C:\Users\tmp_sembene76821\OneDrive - Orange Sonatel\Bureau\AUTO"

dossier= chemin+r"\CA RESET\CURRENT WEEK\inputs"



#Lire tous les fichiers

dataframes=[]

for fichier in os.listdir(dossier):

    if fichier.endswith(".csv"):
        chemin1=os.path.join(dossier,fichier)

        df=pd.read_csv(chemin1,sep="|")


        dataframes.append(df)

df_concat=pd.concat(dataframes, ignore_index=True)

#Renommer les colonnes

df_concat=df_concat.rename(columns={'TARGET_USER_MSISDN':'NUMERO','CHANGED_BY_USER_LOGIN_ID':'EMAIL','CHANGED_BY_USER_NAME':'NOM','CHANGED_BY_USER_LAST_NAME':'PRENOM','CREATED_ON':'DATE ET HEURE','ACTION_TYPE':'OPERATION'})

df_concat=df_concat[['PRENOM','NOM','EMAIL','NUMERO','DATE ET HEURE','OPERATION']]


df_concat.to_excel(chemin+r"\CA RESET\CURRENT WEEK\inputs\tango.xlsx",index=False)

dossier=chemin+r"\CA RESET\CURRENT WEEK"

#Z_susp=glob.glob(dossier + r"\Z_SUSP_IMP*.txt")[0]


#definition des inputs
Case= glob.glob(dossier + r"\Vue Recherche avancée Cases*.xlsx")[0]
Log= chemin+r"\CA RESET\CURRENT WEEK\outputs\INTERMEDIAIRE\logs tango et BO360.xlsx"
BO360=glob.glob(dossier + r"\DEBLO_REINIT*.csv")[0]
login=chemin+r"\CA RESET\CURRENT WEEK\LOGIN\LOGIN BO.xlsx"
ANNUL=glob.glob(dossier + r"\CORRECTION_TRANSACTION*.csv")[0]
CA=chemin+r"\CA RESET\CURRENT WEEK\CA.xlsx"
#BO=r"C:\Users\tmp_sembene76821\OneDrive - Orange Sonatel\Bureau\AUTO\TANGO\CURRENT WEEK\DEBLO_REINIT.xlsx"

print ("Chargement de Log BO360, tango, cases tango et login tango...")

df1= pd.read_excel(Case)
df2= pd.read_excel(login)
df3= pd.read_csv(BO360,sep=";")
#df4= pd.read_csv(ANNUL,sep=";")

#Supprimer doublons dans cases
df1=df1.drop_duplicates(subset=['Numéro mobile'])

df1['Numéro mobile']=pd.to_numeric(df1['Numéro mobile'],errors="coerce")


df1=df1.rename(columns={'Numéro mobile':'NUMERO'})

df1=df1[['NUMERO','Numéro du case']]

#Filtrer actions dans BO360
df3=df3[df3['OPERATION']!='DEBLOCAGE']

#concatener les 2 inputs

Logs=pd.concat([df_concat,df3],ignore_index=True)

Logs.to_excel(chemin+r"\CA RESET\CURRENT WEEK\outputs\INTERMEDIAIRE\logs tango et BO360.xlsx",index=False)

print("Logs TANGO créé avec succés")

#Charger Logs tango

df4=pd.read_excel(Log)

df4=df4.drop_duplicates(subset=['NUMERO','EMAIL'])


#CROISEMENTS

df4['NUMERO']=pd.to_numeric(df4['NUMERO'],errors="coerce")

#login
df2=df2.rename(columns={'LOGIN':'EMAIL'})

#Compatibilité des colonnes

df2['EMAIL']=df2['EMAIL'].astype(str)
Logs['EMAIL']=Logs['EMAIL'].astype(str)

df2['EMAIL']=df2['EMAIL'].str.strip()
Logs['EMAIL']=Logs['EMAIL'].str.strip()



print("Croisement avec les CA")

#Croiser avec CA

df5=pd.read_excel(CA)

df5=df5.rename(columns={'ND Objet CA':'NUMERO'})

df5['NUMERO']=pd.to_numeric(df5['NUMERO'],errors="coerce")

df5=df5[['NUMERO','Date CA']]


Carest=df4.merge(df5, on='NUMERO', how='inner')

#Supprimer differences de casse
df2['EMAIL']=df2['EMAIL'].str.lower()
Carest['EMAIL']=Carest['EMAIL'].str.lower()

print("Croisement avec login")

Actionstango=df2.merge(Carest, on='EMAIL', how='right')

#Actionstango=Actionstango[['STRUCTURE','PRENOM','NOM','EMAIL','NUMERO','DATE ET HEURE','OPERATION']]

#Actionstango.to_excel(r"C:\Users\tmp_sembene76821\OneDrive - Orange Sonatel\Bureau\AUTO\CA RESET\CURRENT WEEK\outputs\INTERMEDIAIRE\actions.xlsx",index=False)

DESC1=Actionstango[Actionstango['STRUCTURE']!='HORS DESC']

with pd.ExcelWriter(chemin+r"\CA RESET\CURRENT WEEK\outputs\CTRLE CA RESET PIN\CAREST.xlsx",engine="xlsxwriter") as writer:

    Actionstango.to_excel(writer, sheet_name="CA RESET",index=False)

    DESC1.to_excel(writer, sheet_name="DESC",index=False)

print("✅Controle CA reset pin ok")


print("Chargement des annulations de transaction")

#Charger annulations

df6=pd.read_csv(ANNUL,sep=";")


df6['EXPEDITEUR']=pd.to_numeric(df6['EXPEDITEUR'],errors="coerce")

df6=df6.rename(columns={'EXPEDITEUR':'NUMERO'})

df6=df6.drop_duplicates(subset=['NUMERO','EMAIL'])


print("Croisement avec login")

df2['EMAIL']=df2['EMAIL'].astype(str).str.strip().str.lower()
df6['EMAIL']=df6['EMAIL'].astype(str).str.strip().str.lower()

annul=df2.merge(df6, on ='EMAIL', how='right')

annul.to_excel(chemin+r"\CA RESET\CURRENT WEEK\outputs\INTERMEDIAIRE\annulations.xlsx",index=False)

#Filtrer sur DESC
DESC=annul[annul['STRUCTURE']!='HORS DESC']

print("Croisement avec cases")

cases=DESC.merge(df1, on ='NUMERO', how='left')

nonconf=cases[cases['Numéro du case'].isna()]

nonconf=nonconf.drop(columns=['VALIDATED','REF ID INITIIATION','INITIATED','REF ID VALIDATION'])



#Croiser avec CA

caannul=DESC.merge(df5, on='NUMERO', how='inner')

caannul['DATE INITIATION']=pd.to_datetime(caannul['DATE INITIATION'],dayfirst=True, errors="raise" )

#format="%d/%m/%Y %H:%M:%S"

caannul['Date CA']=pd.to_datetime(caannul['Date CA'])

caannul['Temps']=caannul['DATE INITIATION'] - caannul['Date CA']

def format_temps(td):
    total_seconds=int(td.total_seconds())
    sign="-" if total_seconds<0 else ''
    total_seconds=abs(total_seconds)
    hours=total_seconds//3600
    mns=(total_seconds%3600)//60
    sec=total_seconds%60
    return f"{sign}{hours}h{mns}m{sec}s"

caannul['Temps de différence']=caannul['Temps'].apply(format_temps)

caannul['CONSTAT']=caannul['Temps'].apply(lambda x: 'NOK' if pd.Timedelta(0) < x < pd.Timedelta(hours=48) else 'OK')

caannul=caannul.drop(columns=['Temps','VALIDATED','REF ID INITIIATION','INITIATED','REF ID VALIDATION'])

NOK=caannul[caannul["CONSTAT"]== "NOK"]

#TCD

#STRUCTURES

pivot_table1=pd.pivot_table(DESC, index='STRUCTURE', aggfunc='size')

pivot_table1 = pivot_table1.reset_index().rename(columns= {0: "Nombre"})

total1= pivot_table1['Nombre'].sum()

totalrow1=pd.DataFrame({'STRUCTURE':['TOTAL'], 'Nombre': [total1]})

pivot_table1=pd.concat([pivot_table1,totalrow1], ignore_index=True)

with pd.ExcelWriter(chemin+r"\CA RESET\CURRENT WEEK\outputs\CTRLE CA RESET PIN\CONTROLE ANNULATION.xlsx",engine="xlsxwriter") as writer:

    cases.to_excel(writer, sheet_name="DESC",index=False)

    nonconf.to_excel(writer, sheet_name="A JUSTIF",index=False)

    caannul.to_excel(writer, sheet_name="CA ANNUL",index=False)

    pivot_table1.to_excel(writer, sheet_name="TCD",index=False)


#FORMATAGE TCD

#STRUCTURE

wb=load_workbook(chemin+r"\CA RESET\CURRENT WEEK\outputs\CTRLE CA RESET PIN\CONTROLE ANNULATION.xlsx")

ws=wb['TCD']

header_fill= PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

header_font= Font(color="FFFFFF", bold=True)

for cell in ws[1]:

    cell.fill= header_fill
    cell.font=header_font

total_fill= PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

total_font= Font(color="FFFFFF", bold=True)

last_row=ws.max_row

for cell in ws[last_row]:

    cell.fill= total_fill
    cell.font=total_font

#BORDURE
thin_side=Side(style='thick')
thin_border=Border(

    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)

max_col=ws.max_column

for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=max_col):

    for cell in row:
        cell.border= thin_border


wb.save(chemin+r"\CA RESET\CURRENT WEEK\outputs\CTRLE CA RESET PIN\CONTROLE ANNULATION.xlsx")


print("✅Controle annulaton ok")

print("Nombre de CA RESET PIN = ",len(Actionstango))
print("Nombre de CA RESET PIN DESC  = ",len(DESC1))
print("Nombre d'annulations' = ",len(annul))
print("Nombre d'annulations DESC= ",len(cases))
print("Nombre d'annulations DESC suspectes = ",len(nonconf))
print("Nombre de CA ANNULATION < 48H = ",len(NOK))

for _ in range(1000000):
    pass
end_time=time.time()

elapsed_time=end_time-start_time

min = elapsed_time // 60
sec= elapsed_time % 60

print(f"Temps d'execution : {int(min)} minutes et {sec:.2f} secondes")